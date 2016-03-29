import openpyxl
from   openpyxl.workbook      import Workbook
from   openpyxl.styles        import Font,Side,Border,Alignment,Color,Style,PatternFill
from   openpyxl.styles.colors import COLOR_INDEX
from   openpyxl.utils         import _get_column_letter
#from   openpyxl.styles.colors import Color
import openpyxl.styles.colors
from   openpyxl.styles.fills  import FILL_SOLID
from   excel.colors           import ColorTable

alignType = {'C':'center','L':'left','R':'right'}
colorType = {'Black':0,'Red':2,'Green':3,'Orange':19}

#---------------------------------------------------------------------
def GetColumnLetter(i):
  return _get_column_letter(i)

def SetColumnWidth(ws,i,width):
  letter = GetColumnLetter(i)
  ws.column_dimensions[letter].width = width 

#---------------------------------------------------------------------
def SetCell(ws,wsRow,wsCol,val,fmt):

  align  = None
  fill   = None
  numFmt = None
  border = None

  c = ws.cell(row=wsRow,column=wsCol)

  for i in fmt:
    #logging.debug(i)

    if (i == 'hAlign'): 
      if (not align): align = Alignment()
      align.horizontal = alignType[fmt[i]]

    if (i == 'vAlign'): 
      if (not align): align = Alignment()
      align.vertical   = alignType[fmt[i]]

    if (i == 'wrap'):
      if (not align): align = Alignment()
      if (fmt[i] == '1'):
        align.wrap_text = 1
      else:
        align.wrap_text = 0

    if (i == 'border'): 
      if (not align): align = Alignment()
      side = Side(style='thin')
      border = Border(left=side,right=side,top=side,bottom=side)

    if (i == 'fill'): 
      color = ColorTable[fmt[i]]
      fill = PatternFill(start_color=color,end_color='FFFFFFFF',fill_type='solid')

    if (i == 'orient'): 
      pass

    if (i == 'bg'):
      fill = PatternFill(start_color='FFEE1111',end_color='FFEE1111',fill_type='solid')

    if (i == 'numFmt'):
      numFmt = fmt[i]

  if (align):
    c.alignment = align.copy()

  if (border):
    c.border = border.copy()

  if (fill):
    c.fill = fill.copy()

  if (numFmt):
    c.number_format = numFmt

  c.value = val

#---------------------------------------------------------------------
def FormatCell(cell,hAlign,vAlign,fmt,fgColor=None,bgColor=None):
  pass

#---------------------------------------------------------------------
def SetCellVal(cell,val,hAlign,vAlign,fmt):
  pass

#---------------------------------------------------------------------
def SetCellBrd(cell,hAlign,vAlign,fmt,value):
  pass

#---------------------------------------------------------------------
def SetCellFmt(cell,hAlign,vAlign,fmt,value):
  if (align == 'C'):
    align = Alignment(horizontal='center',vertical='center')
  elif (align == 'L'):
    align = Alignment(horizontal='left',vertical='center')
  elif (align == 'R'):
    align = Alignment(horizontal='right',vertical='center')
  else:
    align = Alignment(horizontal='right',vertical='center')

  side   = Side(style='thin')
  border = Border(left=side,right=side,top=side,bottom=side)
  #style  = Style(border=border,alignment=align,number_format=fmt)
  #cell.style = style

  if (fmt == 'F'):
    fmt = '0.00'
    cell.number_format = fmt
  cell.alignment     = align.copy()
  cell.border        = border.copy()
  cell.value = value

#---------------------------------------------------------------------
def FlgCell(ws,wsRow,wsCol):

  c = ws.cell(row=wsRow,column=wsCol)

  red = openpyxl.styles.colors.RED
  side   = Side(style='medium',color=red)
  border = Border(left=side,right=side,top=side,bottom=side)
  #style  = Style(border=border)
  #cell.style = style
  c.border = border.copy()

#---------------------------------------------------------------------
def WsCell(ws,wsRow,wsCol,val=None):
  if val is not None:
    return ws.cell(row=wsRow,column=wsCol,value=value)
  else:
    return ws.cell(row=wsRow,column=wsCol)